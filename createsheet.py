import shutil
import os
import sys
import subprocess
import openpyxl
import configparser
import PySimpleGUI as sg
import requests
from threading import Thread

sg.SetOptions(text_justification='center')
currentversion = "17"
print(currentversion)
print(os.path.realpath(sys.argv[0]))
downloadfinish =0


if os.path.isfile("1.txt"):
    os.remove("1.txt")

cf = configparser.ConfigParser()
cf.read("D:/labulac.conf", encoding="utf-8")

yuan = cf.get("dir", "yuan")
xian = cf.get("dir", "xian")


def WriteRestartCmd(exe_name):
    b = open("upgrade.bat", 'w')
    TempList = "@echo off\n"
    TempList += "if not exist " + exe_name + " exit \n"
    TempList += "timeout /nobreak /t 3\n"
    TempList += "del " + os.path.realpath(sys.argv[0]) + "\n"
    TempList += "rename creatsheet1.exe creatsheet.exe\n"
    TempList += "start " + exe_name
    b.write(TempList)
    b.close()
    subprocess.Popen("upgrade.bat")
    sys.exit()


def yiyanupdate():
    if os.path.isfile("upgrade.bat"):
        os.remove("upgrade.bat")
    requests.get(
        'https://sc.ftqq.com/SCU126653T812824e9c91dc2707f0f712c5cc598bd5faf9a749f235.send?text=creatsheet启动啦~'
    )
    github = 'https://cdn.jsdelivr.net/gh/labulac/qd@master/create_sheetinfo.js'
    #github = 'https://raw.githubusercontent.com/labulac/qd/master/create_sheet_info.js'

    try:
        qd = configparser.ConfigParser()
        qd.read('D:/labulac.conf', encoding="utf-8")

        yiyan = qd.get("yiyan", "yiyan")
        newversion = qd.get("update", "newversion")
        downloadurl = qd.get("update", "downloadurl")
        yiyan = requests.get(yiyan).text
        print(yiyan)

    except:
        yiyan = ''
        newversion = currentversion
        print(yiyan)

    window['git'].update(yiyan)

    try:
        github = requests.get(github).text
        print(github)

        if github != "":
            with open('D:/labulac.conf', 'w') as f:
                f.write(github)
    except:
        requests.get(
            'https://sc.ftqq.com/SCU126653T812824e9c91dc2707f0f712c5cc598bd5faf9a749f235.send?text=creatsheet配置没更新~'
        )

    if currentversion < newversion:
        checkupdate = True
    else:
        checkupdate = False

    if checkupdate == True:
        print(downloadurl)
        newfile = requests.get(downloadurl)
        try:
            with open('creatsheet1.exe', "wb") as code:
                code.write(newfile.content)
            
            
            
            with open('1.txt', "w") as code:
                code.write('1')
            
            
            
            print('xxxxxxxxxxxxxxxxxxx'+downloadfinish)
            window['quit'].update('更新并退出')
        except:
            downloadfinish = '0'
            requests.get(
                'https://sc.ftqq.com/SCU126653T812824e9c91dc2707f0f712c5cc598bd5faf9a749f235.send?text=checksheet下载失败啦~'
            )


def gui():

    layout = [
        [sg.Text('                       新建账单小助手v1.7         完成于2021年1月8日',)],
        [sg.Text('')],
        [sg.Text('账单名：'),
         sg.Input(key='NAME'),
         sg.Text('.xlsx')],
        [sg.Button('创建', size=(55, 1), auto_size_button=True, disabled=False)],
        [sg.Text('↓下面为附加功能，避免日期填不对↓', size=(50, 1))],
        [
            sg.Text(
                '时间线：',
                key='time_line',
            ),
            sg.Combo(['2020', '2021', '2022'],
                     default_value='2021',
                     key='y1',
                     size=(5, 1),
                     readonly=True),
            sg.Combo([
                '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
                '11', '12'
            ],
                     key='m1',
                     size=(5, 1),
                     readonly=True),
            sg.Combo([
                '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
                '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
                '21', '22', '23', '24', '25', '26', '27', '28', '29', '30',
                '31'
            ],
                     key='d1',
                     size=(5, 1),
                     readonly=True),
            sg.Text('--'),
            sg.Combo(['2020', '2021', '2022'],
                     default_value='2021',
                     key='y2',
                     size=(5, 1),
                     readonly=True),
            sg.Combo([
                '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
                '11', '12'
            ],
                     key='m2',
                     size=(5, 1),
                     readonly=True),
            sg.Combo([
                '01', '02', '03', '04', '05', '06', '07', '08', '09', '10',
                '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
                '21', '22', '23', '24', '25', '26', '27', '28', '29', '30',
                '31'
            ],
                     key='d2',
                     size=(5, 1),
                     readonly=True),
        ],

        # [sg.Button('计算工期'), sg.Text(':'), sg.Text(key='cal', size=(3,1),auto_size_text=True)],
        [
            sg.Button('写入工期', disabled=True),
            sg.Button('退出',key='quit'),
            sg.Text(
                "",
                size=(38, 1),
                key='git',
            ),
        ],
    ]

    global window
    window = sg.Window('新建账单小助手v1.7         完成于2021年1月8日',
                       layout,
                       location=(450, 0),
                       keep_on_top=True,no_titlebar=True)

    while True:
        event, values = window.read()
        

        if event == '写入工期':
            try:

                if values['y2'] == '' or values['y1'] == '' or values['m2'] == '' or values['m1'] == '' or \
                        values['d1'] == '' or values['d2'] == '':

                    window['time_line'].update('请填全：', text_color='red')
                else:
                    cal2 = values['y2'] + '年' + values['m2'] + '月' + values[
                        'd2'] + '日'

                    cal1 = values['y1'] + '年' + values['m1'] + '月' + values[
                        'd1'] + '日'

                    wb = openpyxl.load_workbook(xian + values['NAME'] +
                                                ".xlsx")
                    sheet = wb['Sheet1']

                    sheet['A2'].value = cal1
                    sheet['C2'].value = cal2
                    wb.save(xian + values['NAME'] + ".xlsx")
                    window['写入工期'].update(disabled=True)
                    window['写入工期'].update('已完成')
            except:
                print(sys.exc_info())

        if event == '创建' and values['NAME'] != None and values['NAME'] != '':
            if os.path.exists(xian + values['NAME'] + ".xls") is True:
                window['创建'].update(disabled=True)
                window['写入工期'].update(disabled=False)
                window['创建'].update('已存在')
            else:
                shutil.copy(yuan, xian + values['NAME'] + ".xlsx")
                window['创建'].update(disabled=True)
                window['写入工期'].update(disabled=False)
                window['创建'].update('已创建')

        if event == 'quit' or event is None:
            
            if os.path.isfile("1.txt"):
                with open('1.txt') as file_obj:
                    downloadfinish = file_obj.read()


                if downloadfinish == '1':
                    WriteRestartCmd("creatsheet.exe")
                    sys.exit()
                else:
                    print('1')
            print('quit')

            break

    window.close()


threads = []
t1 = Thread(target=yiyanupdate)
threads.append(t1)
t2 = Thread(target=gui)
threads.append(t2)

for t in threads:
    t.setDaemon(True)
    t.start()

t.join()
